# bulk_operations/services/template_generator.py
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import json


class ExcelTemplateGenerator:
    """
    Service for generating Excel templates with data validation, sample data, and instructions.
    """
    
    HEADER_STYLE = {
        'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
        'font': Font(color='FFFFFF', bold=True),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    SAMPLE_STYLE = {
        'fill': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid'),
        'font': Font(italic=True),
        'alignment': Alignment(horizontal='left', vertical='center'),
    }

    @classmethod
    def generate_excel_template(cls, template_type: str) -> Workbook:
        """Generate an Excel template with validation and sample data."""
        wb = Workbook()
        
        # Create Instructions sheet first
        instructions_sheet = wb.create_sheet('Instructions', 0)
        
        # Create Template sheet
        template_sheet = wb.create_sheet('Template', 1)
        wb.active = template_sheet  # Make Template sheet active
        
        # Get template definition
        template_def = cls.get_template_definition(template_type)
        fields = template_def['fields']
        field_types = cls.get_field_types(template_type)
        
        # Add instructions
        cls._add_instructions(instructions_sheet, template_type, fields, field_types)
        
        # Add headers and styling to Template sheet
        for col, field in enumerate(fields, start=1):
            # Add header
            cell = template_sheet.cell(row=1, column=col, value=field)
            
            # Add required indicator
            if field_types[field].get('required'):
                cell.value = f"{field} *"
            
            # Apply header styling
            for style_attr, style_value in cls.HEADER_STYLE.items():
                setattr(cell, style_attr, style_value)
            
            # Add field description as comment
            field_info = field_types[field]
            comment_text = f"Type: {field_info['type']}\n"
            if field_info.get('description'):
                comment_text += f"Description: {field_info['description']}\n"
            if field_info.get('required'):
                comment_text += "Required: Yes\n"
            if field_info.get('choices'):
                choices = [choice[0] for choice in field_info['choices']]
                comment_text += f"Valid values: {', '.join(choices)}\n"
            if field_info.get('max_length'):
                comment_text += f"Maximum length: {field_info['max_length']}\n"
            
            comment = Comment(comment_text, "Template Generator")
            cell.comment = comment
            
            # Set column width based on content and field type
            width = max(len(str(cell.value)) + 5, 15)
            if field_info['type'] in ['string', 'email']:
                width = min(max(width, field_info.get('max_length', 50) // 5), 50)
            template_sheet.column_dimensions[get_column_letter(col)].width = width
            
            # Add data validation
            cls._add_field_validation(template_sheet, col, field_info)
        
        # Add sample data
        sample_data = cls._get_sample_data(template_type)
        if sample_data:
            for row, record in enumerate(sample_data, start=2):
                for col, field in enumerate(fields, start=1):
                    cell = template_sheet.cell(row=row, column=col, value=record.get(field, ''))
                    for style_attr, style_value in cls.SAMPLE_STYLE.items():
                        setattr(cell, style_attr, style_value)
        
        # Freeze header row
        template_sheet.freeze_panes = 'A2'
        
        return wb

    @classmethod
    def _add_instructions(cls, ws, template_type: str, fields: List[str], field_types: Dict):
        """Add instructions sheet with detailed information."""
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = f"{template_type.replace('_', ' ').title()} Template Instructions"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add general instructions
        ws['A3'] = "General Instructions:"
        ws['A4'] = "1. Required fields are marked with an asterisk (*) in the template."
        ws['A5'] = "2. Use the sample data as a reference for the expected format."
        ws['A6'] = "3. Data validation is enabled to help prevent errors."
        ws['A7'] = "4. Hover over column headers to see detailed field descriptions."
        
        # Add field details
        ws['A9'] = "Field Details:"
        ws['A10'] = "Field Name"
        ws['B10'] = "Type"
        ws['C10'] = "Required"
        ws['D10'] = "Description"
        
        for style_attr, style_value in cls.HEADER_STYLE.items():
            for cell in ws['10']:
                setattr(cell, style_attr, style_value)
        
        for i, field in enumerate(fields, start=11):
            field_info = field_types[field]
            ws[f'A{i}'] = field
            ws[f'B{i}'] = field_info['type']
            ws[f'C{i}'] = 'Yes' if field_info.get('required') else 'No'
            ws[f'D{i}'] = field_info.get('description', '')
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50

    @classmethod
    def _add_field_validation(cls, ws, col: int, field_info: Dict):
        """Add data validation rules based on field type."""
        col_letter = get_column_letter(col)
        validation = None
        
        if field_info['type'] == 'choice':
            choices = [choice[0] for choice in field_info.get('choices', [])]
            if choices:
                validation = DataValidation(
                    type='list',
                    formula1=f'"{",".join(choices)}"',
                    allow_blank=not field_info.get('required', False)
                )
        
        elif field_info['type'] == 'integer':
            validation = DataValidation(
                type='whole',
                operator='greaterThanOrEqual',
                formula1='0',
                allow_blank=not field_info.get('required', False)
            )
        
        elif field_info['type'] == 'decimal':
            validation = DataValidation(
                type='decimal',
                operator='greaterThanOrEqual',
                formula1='0',
                allow_blank=not field_info.get('required', False)
            )
        
        elif field_info['type'] == 'email':
            # Basic email format validation
            validation = DataValidation(
                type='custom',
                formula1='=AND(ISNUMBER(FIND("@",{col}2)),ISNUMBER(FIND(".",{col}2,FIND("@",{col}2))))'
                .format(col=col_letter),
                allow_blank=not field_info.get('required', False)
            )
        
        if validation:
            validation.error = 'Invalid value'
            validation.errorTitle = 'Validation Error'
            validation.prompt = field_info.get('description', '')
            validation.promptTitle = field_info.get('name', field_info['type'].title())
            ws.add_data_validation(validation)
            validation.add(f'{col_letter}2:{col_letter}1048576')

    @classmethod
    def _get_sample_data(cls, template_type: str) -> List[Dict]:
        """Get sample data for the template type."""
        if template_type == 'customers':
            return [
                {
                    'company_name': 'Acme Corp',
                    'legal_business_name': 'Acme Corporation Inc.',
                    'email': 'contact@acme.com',
                    'phone': '555-0123',
                    'address': '123 Business St',
                    'city': 'Springfield',
                    'state': 'IL',
                    'zip': '62701',
                    'country': 'USA',
                    'business_type': 'Manufacturing'
                }
            ]
        elif template_type == 'orders':
            return [
                {
                    'transaction_id': 1001,
                    'customer': 1,
                    'reference_number': 'ORD-2024-001',
                    'ship_to_name': 'John Smith',
                    'ship_to_company': 'Acme Corp',
                    'ship_to_address': '456 Shipping Lane',
                    'ship_to_city': 'Chicago',
                    'ship_to_state': 'IL',
                    'ship_to_zip': '60601',
                    'weight_lb': 25.5,
                    'line_items': 3,
                    'sku_quantity': json.dumps({'SKU001': 5, 'SKU002': 3}),
                    'status': 'draft',
                    'priority': 'medium'
                }
            ]
        elif template_type == 'products':
            return [
                {
                    'sku': 'PROD-001',
                    'customer': 1,
                    'labeling_unit_1': 'Box',
                    'labeling_quantity_1': 10,
                    'labeling_unit_2': 'Case',
                    'labeling_quantity_2': 5,
                    'labeling_unit_3': 'Pallet',
                    'labeling_quantity_3': 40,
                    'labeling_unit_4': '',
                    'labeling_quantity_4': None,
                    'labeling_unit_5': '',
                    'labeling_quantity_5': None
                }
            ]
        elif template_type == 'services':
            return [
                {
                    'service_name': 'Custom Packaging',
                    'description': 'Professional packaging service with custom materials',
                    'charge_type': 'quantity'
                }
            ]
        elif template_type == 'materials':
            return [
                {
                    'name': 'Cardboard Box - Large',
                    'description': 'Heavy-duty corrugated cardboard box, 24x24x24 inches',
                    'unit_price': 12.99
                }
            ]
        elif template_type == 'inserts':
            return [
                {
                    'sku': 'INS-001',
                    'insert_name': 'Product Manual',
                    'insert_quantity': 1,
                    'customer': 1
                }
            ]
        elif template_type == 'cad_shipping':
            return [
                {
                    'transaction': 1001,
                    'customer': 1,
                    'service_code_description': 'Express Overnight',
                    'ship_to_name': 'Jane Smith',
                    'ship_to_address_1': '789 Delivery Ave',
                    'ship_to_address_2': 'Suite 101',
                    'ship_to_city': 'Toronto',
                    'ship_to_state': 'ON',
                    'ship_to_postal_code': 'M5V 2T6',
                    'ship_to_country': 'CA',
                    'tracking_number': 'CAD123456789',
                    'pre_tax_shipping_charge': 45.99,
                    'tax1type': 'HST',
                    'tax1amount': 5.98,
                    'fuel_surcharge': 2.50,
                    'reference': 'REF-001',
                    'weight': 15.5,
                    'box_length': 12,
                    'box_width': 8,
                    'box_height': 6,
                    'carrier': 'Canada Post'
                }
            ]
        elif template_type == 'us_shipping':
            return [
                {
                    'transaction': 1002,
                    'customer': 1,
                    'ship_date': '2024-02-15',
                    'ship_to_name': 'Bob Johnson',
                    'ship_to_address_1': '321 Delivery Blvd',
                    'ship_to_address_2': 'Floor 3',
                    'ship_to_city': 'New York',
                    'ship_to_state': 'NY',
                    'ship_to_zip': '10001',
                    'ship_to_country_code': 'US',
                    'tracking_number': 'US987654321',
                    'service_name': 'Ground',
                    'weight_lbs': 18.5,
                    'length_in': 15,
                    'width_in': 10,
                    'height_in': 8,
                    'base_chg': 35.99,
                    'carrier_peak_charge': 5.00,
                    'wizmo_peak_charge': 2.50,
                    'accessorial_charges': 7.50,
                    'rate': 50.99,
                    'current_status': 'in_transit',
                    'delivery_status': 'pending'
                }
            ]
        return []

    @classmethod
    def get_template_definition(cls, template_type: str) -> Dict[str, Any]:
        """Get template definition including fields and required fields."""
        fields = cls._get_field_definitions(template_type)
        return {
            'fields': list(fields.keys()),
            'required_fields': [f for f, v in fields.items() if v.get('required', False)]
        }

    @classmethod
    def get_field_types(cls, template_type: str) -> Dict[str, Dict[str, Any]]:
        """Get field types and metadata for a template."""
        return cls._get_field_definitions(template_type)

    @classmethod
    def _get_field_definitions(cls, template_type: str) -> Dict[str, Dict[str, Any]]:
        """Get complete field definitions for a template type."""
        if template_type == 'customers':
            return {
                'company_name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Official company name as it appears on legal documents',
                    'max_length': 255
                },
                'legal_business_name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Legal registered business name',
                    'max_length': 255
                },
                'email': {
                    'type': 'email',
                    'required': True,
                    'description': 'Primary contact email address'
                },
                'phone': {
                    'type': 'string',
                    'required': False,
                    'description': 'Primary contact phone number',
                    'max_length': 20
                },
                'address': {
                    'type': 'string',
                    'required': False,
                    'description': 'Street address',
                    'max_length': 255
                },
                'city': {
                    'type': 'string',
                    'required': False,
                    'description': 'City name',
                    'max_length': 100
                },
                'state': {
                    'type': 'string',
                    'required': False,
                    'description': 'State/Province',
                    'max_length': 100
                },
                'zip': {
                    'type': 'string',
                    'required': False,
                    'description': 'ZIP/Postal code',
                    'max_length': 20
                },
                'country': {
                    'type': 'string',
                    'required': False,
                    'description': 'Country name',
                    'max_length': 100
                },
                'business_type': {
                    'type': 'string',
                    'required': False,
                    'description': 'Type of business',
                    'max_length': 50
                }
            }
        elif template_type == 'orders':
            return {
                'transaction_id': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Unique identifier for the order transaction'
                },
                'customer': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Customer ID associated with the order'
                },
                'reference_number': {
                    'type': 'string',
                    'required': True,
                    'description': 'Order reference number',
                    'max_length': 50
                },
                'ship_to_name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Name of the recipient',
                    'max_length': 255
                },
                'ship_to_company': {
                    'type': 'string',
                    'required': False,
                    'description': 'Company name for shipping',
                    'max_length': 255
                },
                'ship_to_address': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping street address',
                    'max_length': 255
                },
                'ship_to_address2': {
                    'type': 'string',
                    'required': False,
                    'description': 'Additional shipping address details',
                    'max_length': 255
                },
                'ship_to_city': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping city',
                    'max_length': 100
                },
                'ship_to_state': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping state/province',
                    'max_length': 100
                },
                'ship_to_zip': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping ZIP/postal code',
                    'max_length': 20
                },
                'ship_to_country': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipping country',
                    'max_length': 100
                },
                'weight_lb': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Order weight in pounds'
                },
                'line_items': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Number of line items in the order'
                },
                'sku_quantity': {
                    'type': 'string',
                    'required': False,
                    'description': 'JSON object mapping SKUs to quantities (e.g., {"SKU001": 5, "SKU002": 3})',
                    'max_length': 1000
                },
                'total_item_qty': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Total quantity of all items'
                },
                'volume_cuft': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Order volume in cubic feet'
                },
                'packages': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Number of packages in the order'
                },
                'notes': {
                    'type': 'string',
                    'required': False,
                    'description': 'Additional notes or comments',
                    'max_length': 1000
                },
                'carrier': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipping carrier',
                    'max_length': 50
                },
                'status': {
                    'type': 'choice',
                    'required': False,
                    'description': 'Order status',
                    'choices': [
                        ('draft', 'Draft'),
                        ('submitted', 'Submitted'),
                        ('shipped', 'Shipped'),
                        ('delivered', 'Delivered'),
                        ('cancelled', 'Cancelled')
                    ],
                    'default': 'draft'
                },
                'priority': {
                    'type': 'choice',
                    'required': False,
                    'description': 'Order priority level',
                    'choices': [
                        ('low', 'Low'),
                        ('medium', 'Medium'),
                        ('high', 'High')
                    ],
                    'default': 'medium'
                }
            }
        elif template_type == 'products':
            return {
                'sku': {
                    'type': 'string',
                    'required': True,
                    'description': 'Stock Keeping Unit - unique product identifier',
                    'max_length': 50
                },
                'customer': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Customer ID associated with the product'
                },
                'labeling_unit_1': {
                    'type': 'string',
                    'required': False,
                    'description': 'Primary labeling unit',
                    'max_length': 50
                },
                'labeling_quantity_1': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Quantity for primary labeling unit'
                },
                'labeling_unit_2': {
                    'type': 'string',
                    'required': False,
                    'description': 'Secondary labeling unit',
                    'max_length': 50
                },
                'labeling_quantity_2': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Quantity for secondary labeling unit'
                },
                'labeling_unit_3': {
                    'type': 'string',
                    'required': False,
                    'description': 'Tertiary labeling unit',
                    'max_length': 50
                },
                'labeling_quantity_3': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Quantity for tertiary labeling unit'
                },
                'labeling_unit_4': {
                    'type': 'string',
                    'required': False,
                    'description': 'Quaternary labeling unit',
                    'max_length': 50
                },
                'labeling_quantity_4': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Quantity for quaternary labeling unit'
                },
                'labeling_unit_5': {
                    'type': 'string',
                    'required': False,
                    'description': 'Quinary labeling unit',
                    'max_length': 50
                },
                'labeling_quantity_5': {
                    'type': 'integer',
                    'required': False,
                    'description': 'Quantity for quinary labeling unit'
                }
            }
        elif template_type == 'services':
            return {
                'service_name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Name of the service',
                    'max_length': 255
                },
                'description': {
                    'type': 'string',
                    'required': False,
                    'description': 'Detailed description of the service',
                    'max_length': 1000
                },
                'charge_type': {
                    'type': 'choice',
                    'required': True,
                    'description': 'Type of service charge',
                    'choices': [
                        ('single', 'Single Charge'),
                        ('quantity', 'Quantity Based')
                    ],
                    'default': 'quantity'
                }
            }
        elif template_type == 'materials':
            return {
                'name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Name of the material',
                    'max_length': 255
                },
                'description': {
                    'type': 'string',
                    'required': False,
                    'description': 'Detailed description of the material',
                    'max_length': 1000
                },
                'unit_price': {
                    'type': 'decimal',
                    'required': True,
                    'description': 'Price per unit of material'
                }
            }
        elif template_type == 'inserts':
            return {
                'sku': {
                    'type': 'string',
                    'required': True,
                    'description': 'Stock Keeping Unit - unique insert identifier',
                    'max_length': 50
                },
                'insert_name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Name of the insert',
                    'max_length': 255
                },
                'insert_quantity': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Quantity of inserts'
                },
                'customer': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Customer ID associated with the insert'
                }
            }
        elif template_type == 'cad_shipping':
            return {
                'transaction': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Order ID for the shipment'
                },
                'customer': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Customer ID associated with the shipment'
                },
                'service_code_description': {
                    'type': 'string',
                    'required': False,
                    'description': 'Description of shipping service',
                    'max_length': 255
                },
                'ship_to_name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Name of the recipient',
                    'max_length': 255
                },
                'ship_to_address_1': {
                    'type': 'string',
                    'required': True,
                    'description': 'Primary shipping address',
                    'max_length': 255
                },
                'ship_to_address_2': {
                    'type': 'string',
                    'required': False,
                    'description': 'Secondary shipping address',
                    'max_length': 255
                },
                'shiptoaddress3': {
                    'type': 'string',
                    'required': False,
                    'description': 'Additional shipping address',
                    'max_length': 255
                },
                'ship_to_city': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping city',
                    'max_length': 100
                },
                'ship_to_state': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping province',
                    'max_length': 100
                },
                'ship_to_country': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipping country (defaults to CA)',
                    'max_length': 100,
                    'default': 'CA'
                },
                'ship_to_postal_code': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping postal code',
                    'max_length': 20
                },
                'tracking_number': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipment tracking number',
                    'max_length': 100
                },
                'pre_tax_shipping_charge': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Shipping charge before taxes'
                },
                'tax1type': {
                    'type': 'string',
                    'required': False,
                    'description': 'Type of first tax',
                    'max_length': 50
                },
                'tax1amount': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Amount of first tax'
                },
                'tax2type': {
                    'type': 'string',
                    'required': False,
                    'description': 'Type of second tax',
                    'max_length': 50
                },
                'tax2amount': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Amount of second tax'
                },
                'tax3type': {
                    'type': 'string',
                    'required': False,
                    'description': 'Type of third tax',
                    'max_length': 50
                },
                'tax3amount': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Amount of third tax'
                },
                'fuel_surcharge': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Fuel surcharge amount'
                },
                'reference': {
                    'type': 'string',
                    'required': False,
                    'description': 'Reference number',
                    'max_length': 100
                },
                'weight': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Shipment weight'
                },
                'gross_weight': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Gross weight of shipment'
                },
                'box_length': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Length of shipping box'
                },
                'box_width': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Width of shipping box'
                },
                'box_height': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Height of shipping box'
                },
                'box_name': {
                    'type': 'string',
                    'required': False,
                    'description': 'Name or identifier of the box',
                    'max_length': 100
                },
                'carrier': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipping carrier',
                    'max_length': 100
                }
            }
        elif template_type == 'us_shipping':
            return {
                'transaction': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Order ID for the shipment'
                },
                'customer': {
                    'type': 'integer',
                    'required': True,
                    'description': 'Customer ID associated with the shipment'
                },
                'ship_date': {
                    'type': 'string',
                    'required': False,
                    'description': 'Date of shipment (YYYY-MM-DD)',
                    'max_length': 10
                },
                'ship_to_name': {
                    'type': 'string',
                    'required': True,
                    'description': 'Name of the recipient',
                    'max_length': 255
                },
                'ship_to_address_1': {
                    'type': 'string',
                    'required': True,
                    'description': 'Primary shipping address',
                    'max_length': 255
                },
                'ship_to_address_2': {
                    'type': 'string',
                    'required': False,
                    'description': 'Secondary shipping address',
                    'max_length': 255
                },
                'ship_to_city': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping city',
                    'max_length': 100
                },
                'ship_to_state': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping state',
                    'max_length': 100
                },
                'ship_to_zip': {
                    'type': 'string',
                    'required': True,
                    'description': 'Shipping ZIP code',
                    'max_length': 20
                },
                'ship_to_country_code': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipping country code (defaults to US)',
                    'max_length': 2,
                    'default': 'US'
                },
                'tracking_number': {
                    'type': 'string',
                    'required': False,
                    'description': 'Shipment tracking number',
                    'max_length': 100
                },
                'service_name': {
                    'type': 'string',
                    'required': False,
                    'description': 'Name of shipping service',
                    'max_length': 100
                },
                'weight_lbs': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Weight in pounds'
                },
                'length_in': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Length in inches'
                },
                'width_in': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Width in inches'
                },
                'height_in': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Height in inches'
                },
                'base_chg': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Base shipping charge'
                },
                'carrier_peak_charge': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Carrier peak season charge'
                },
                'wizmo_peak_charge': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Wizmo peak season charge'
                },
                'accessorial_charges': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Additional handling charges'
                },
                'rate': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Total shipping rate'
                },
                'hst': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Harmonized Sales Tax'
                },
                'gst': {
                    'type': 'decimal',
                    'required': False,
                    'description': 'Goods and Services Tax'
                },
                'current_status': {
                    'type': 'choice',
                    'required': False,
                    'description': 'Current shipment status',
                    'choices': [
                        ('pending', 'Pending'),
                        ('in_transit', 'In Transit'),
                        ('delivered', 'Delivered'),
                        ('exception', 'Exception')
                    ],
                    'default': 'pending'
                },
                'delivery_status': {
                    'type': 'choice',
                    'required': False,
                    'description': 'Delivery status',
                    'choices': [
                        ('pending', 'Pending'),
                        ('attempted', 'Attempted'),
                        ('delivered', 'Delivered'),
                        ('failed', 'Failed')
                    ],
                    'default': 'pending'
                }
            }
        return {}