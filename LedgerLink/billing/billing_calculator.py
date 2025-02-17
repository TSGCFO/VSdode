# billing_calculator.py

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Optional, Union
import json
from django.core.exceptions import ValidationError
from django.db.models import Q
import logging
from openpyxl import Workbook
from io import BytesIO

from orders.models import Order
from customers.models import Customer
from services.models import Service
from rules.models import Rule, RuleGroup
from customer_services.models import CustomerService
from products.models import Product

logger = logging.getLogger(__name__)


@dataclass
class ServiceCost:
    """
    Represents the cost details of a specific service.

    This class holds information about a service's identifier, name, and
    the monetary amount associated with it. It can be used to track
    service costs in billing systems or for general financial record-keeping.

    :ivar service_id: Unique identifier for the service.
    :type service_id: int
    :ivar service_name: Human-readable name of the service.
    :type service_name: str
    :ivar amount: Cost amount associated with the service.
    :type amount: Decimal
    """
    service_id: int
    service_name: str
    amount: Decimal


@dataclass
class OrderCost:
    """
    Represents the cost details of an order.

    This class is a data structure to encapsulate the cost-related components
    of an order. It includes the order's identifier, a list of associated 
    service costs, and the total amount for the order. It is intended to 
    facilitate clear organization and management of order cost data.

    :ivar order_id: Unique identifier for the order.
    :type order_id: int
    :ivar service_costs: List of service costs related to the order.
    :type service_costs: List[ServiceCost]
    :ivar total_amount: Total cost of the order, including all services.
    :type total_amount: Decimal
    """
    order_id: int
    service_costs: List[ServiceCost] = field(default_factory=list)
    total_amount: Decimal = Decimal('0')


@dataclass
class BillingReport:
    """
    Represents a billing report for a specific customer over a given time period.

    This class encapsulates the billing details, including costs of individual 
    orders, service-specific totals, and the overall total amount billed. It is 
    designed to provide a structured summary of billing information for reporting 
    and analysis purposes.

    :ivar customer_id: ID of the customer the billing report belongs to.
    :type customer_id: int
    :ivar start_date: Start date of the billing period.
    :type start_date: datetime
    :ivar end_date: End date of the billing period.
    :type end_date: datetime
    :ivar order_costs: List of individual order costs for the billing period.
    :type order_costs: List[OrderCost]
    :ivar service_totals: Aggregated costs by service type, keyed by
        service ID.
    :type service_totals: Dict[int, Decimal]
    :ivar total_amount: Total amount billed for the customer in the
        given period.
    :type total_amount: Decimal
    :ivar report_data: Dictionary containing the report data in a format
        suitable for serialization.
    :type report_data: Dict
    """
    customer_id: int
    start_date: datetime
    end_date: datetime
    order_costs: List[OrderCost] = field(default_factory=list)
    service_totals: Dict[int, Decimal] = field(default_factory=dict)
    total_amount: Decimal = Decimal('0')
    report_data: Dict = field(default_factory=dict)


def normalize_sku(sku: str) -> str:
    """
    Normalize a given SKU (Stock Keeping Unit) by removing extra spaces,
    converting it to uppercase, and ensuring the result is a valid string.
    The function handles empty or invalid inputs gracefully by returning an
    empty string.

    :param sku: The SKU to normalize.
    :type sku: str
    :return: A normalized SKU where extra spaces are removed, and all letters
        are in uppercase. If the input is invalid or empty, returns an
        empty string.
    :rtype: str
    """
    try:
        if not sku:
            return ''
        # Remove extra spaces and convert to uppercase
        return ''.join(str(sku).split()).upper()
    except (AttributeError, TypeError):
        return ''


def convert_sku_format(sku_data) -> Dict:
    """
    Converts SKU data into a normalized dictionary format. The input SKU data can be
    provided as a JSON-encoded string or a list of dictionaries. Each dictionary should 
    contain two keys: 'sku' and 'quantity'. The function normalizes SKUs and aggregates 
    quantities for duplicate SKUs. Invalid entries are logged and excluded from the result.

    :param sku_data: Input SKU data that can be of type `str` (JSON-encoded) 
        or a `list` of dictionaries, where each dictionary must include 
        'sku' and 'quantity' keys.
    :rtype: Dict
    :return: A dictionary where keys are normalized SKUs and values are 
        the aggregated quantities. Returns an empty dictionary if the input 
        is invalid or contains errors.
    """
    try:
        if isinstance(sku_data, str):
            sku_data = json.loads(sku_data)

        if not isinstance(sku_data, list):
            logger.error(f"SKU data must be a list, got {type(sku_data)}")
            return {}

        sku_dict = {}
        for item in sku_data:
            if not isinstance(item, dict):
                logger.error(f"Each SKU item must be a dictionary, got {type(item)}")
                continue

            if 'sku' not in item or 'quantity' not in item:
                logger.error("SKU item missing required fields 'sku' or 'quantity'")
                continue

            # Normalize SKU format
            sku = normalize_sku(str(item['sku']))
            if not sku:
                logger.error("SKU cannot be empty")
                continue

            try:
                quantity = float(item['quantity'])
            except (TypeError, ValueError):
                logger.error(f"Invalid quantity for SKU {sku}: {item['quantity']}")
                continue

            if quantity <= 0:
                logger.error(f"Invalid quantity {quantity} for SKU {sku}")
                continue

            # If the same SKU appears multiple times (with different formats),
            # add the quantities together
            sku_dict[sku] = sku_dict.get(sku, 0) + quantity

        return sku_dict
    except (json.JSONDecodeError, TypeError, KeyError) as e:
        logger.error(f"Error converting SKU format: {str(e)}")
        return {}


def validate_sku_quantity(sku_data) -> bool:
    """
    Validates the SKU data to ensure that each SKU is a non-empty string and each associated
    quantity is a positive numeric value. Allows input in stringified JSON or list format and 
    processes it to dictionary form for validation.

    :param sku_data: SKU data to be validated. It can be a stringified JSON or a list of SKUs 
        and their quantities.
    :type sku_data: Union[str, list]

    :return: True if the SKU data is valid, otherwise False.
    :rtype: bool
    """
    try:
        if isinstance(sku_data, str):
            sku_data = json.loads(sku_data)

        if not isinstance(sku_data, list):
            return False

        # Convert to dictionary
        sku_dict = convert_sku_format(sku_data)
        if not sku_dict:
            return False

        # Validate the converted dictionary
        for sku, quantity in sku_dict.items():
            if not isinstance(sku, str) or not sku.strip():
                return False

            try:
                qty = float(quantity)
                if qty <= 0:
                    return False
            except (TypeError, ValueError):
                return False

        return True
    except Exception as e:
        logger.error(f"Error validating SKU quantity: {str(e)}")
        return False


class RuleEvaluator:
    """
    Responsible for evaluating rules and rule groups in the given context.

    This class provides functionality to evaluate individual rules and rule groups
    based on their defined fields, operators, and associated conditions. It performs
    validation and comparison operations on order objects and their relevant attributes,
    while supporting both numeric and string fields, as well as complex field types
    like SKU quantities.

    :ivar logger: Logger instance to log warnings and errors encountered during rule
                  evaluation.
    :type logger: logging.Logger
    """
    @staticmethod
    def evaluate_rule(rule: Rule, order: Order) -> bool:
        try:
            field_value = getattr(order, rule.field, None)
            if field_value is None:
                logger.warning(f"Field {rule.field} not found in order {order.transaction_id}")
                return False

            values = rule.get_values_as_list()

            # Handle numeric fields
            numeric_fields = ['weight_lb', 'line_items', 'total_item_qty', 'volume_cuft', 'packages']
            if rule.field in numeric_fields:
                try:
                    field_value = float(field_value) if field_value is not None else 0
                    value = float(values[0]) if values else 0

                    if rule.operator == 'gt':
                        return field_value > value
                    elif rule.operator == 'lt':
                        return field_value < value
                    elif rule.operator == 'eq':
                        return field_value == value
                    elif rule.operator == 'ne':
                        return field_value != value
                    elif rule.operator == 'ge':
                        return field_value >= value
                    elif rule.operator == 'le':
                        return field_value <= value
                except (ValueError, TypeError):
                    logger.error(f"Error converting numeric values for field {rule.field}")
                    return False

            # Handle string fields
            string_fields = ['reference_number', 'ship_to_name', 'ship_to_company',
                             'ship_to_city', 'ship_to_state', 'ship_to_country',
                             'carrier', 'notes']
            if rule.field in string_fields:
                field_value = str(field_value) if field_value is not None else ''

                if rule.operator == 'eq':
                    return field_value == values[0]
                elif rule.operator == 'ne':
                    return field_value != values[0]
                elif rule.operator == 'in':
                    return field_value in values
                elif rule.operator == 'ni':
                    return field_value not in values
                elif rule.operator == 'contains':
                    return any(v in field_value for v in values)
                elif rule.operator == 'ncontains':
                    return not any(v in field_value for v in values)
                elif rule.operator == 'startswith':
                    return any(field_value.startswith(v) for v in values)
                elif rule.operator == 'endswith':
                    return any(field_value.endswith(v) for v in values)

            # Handle SKU quantity
            if rule.field == 'sku_quantity':
                if field_value is None:
                    return False

                try:
                    if isinstance(field_value, str):
                        field_value = json.loads(field_value)

                    if not validate_sku_quantity(field_value):
                        return False

                    # Extract SKUs for comparison
                    sku_dict = convert_sku_format(field_value)
                    # Normalize SKUs for comparison
                    skus = [normalize_sku(sku) for sku in sku_dict.keys()]

                    if rule.operator == 'contains':
                        return any(normalize_sku(v) in skus for v in values)
                    elif rule.operator == 'ncontains':
                        return not any(normalize_sku(v) in skus for v in values)
                    elif rule.operator == 'in':
                        return any(normalize_sku(v) in str(sku_dict) for v in values)
                    elif rule.operator == 'ni':
                        return not any(normalize_sku(v) in str(sku_dict) for v in values)
                except (json.JSONDecodeError, AttributeError):
                    logger.error(f"Error processing SKU quantity")
                    return False

            logger.warning(f"Unhandled field {rule.field} or operator {rule.operator}")
            return False

        except Exception as e:
            logger.error(f"Error evaluating rule: {str(e)}")
            return False

    @staticmethod
    def evaluate_rule_group(rule_group: RuleGroup, order: Order) -> bool:
        try:
            rules = rule_group.rules.all()
            if not rules:
                logger.warning(f"No rules found in rule group {rule_group.id}")
                return False

            results = [RuleEvaluator.evaluate_rule(rule, order) for rule in rules]

            if rule_group.logic_operator == 'AND':
                return all(results)
            elif rule_group.logic_operator == 'OR':
                return any(results)
            elif rule_group.logic_operator == 'NOT':
                return not any(results)
            elif rule_group.logic_operator == 'XOR':
                return sum(results) == 1
            elif rule_group.logic_operator == 'NAND':
                return not all(results)
            elif rule_group.logic_operator == 'NOR':
                return not any(results)

            logger.warning(f"Unknown logic operator {rule_group.logic_operator}")
            return False

        except Exception as e:
            logger.error(f"Error evaluating rule group: {str(e)}")
            return False


class BillingCalculator:
    """
    The BillingCalculator class is responsible for calculating costs associated with a customer's
    services based on various parameters such as order details, service charge types, and assigned SKUs.
    It encapsulates the logic for validating inputs, identifying applicable services, and generating a
    comprehensive billing report for the specified time range.

    This class makes use of intricate business logic for quantity-based services, including detailed
    matching of SKUs, calculation of full cases, and the handling of other specific service types such
    as 'Pick Cost', 'Case Pick', and 'SKU Cost'. It also facilitates logging for auditing and debugging
    purposes. Services without SKUs are calculated based on preset base prices, with additional support
    for uncommon cases.

    :ivar customer_id: The unique ID representing the customer.
    :type customer_id: int
    :ivar start_date: The starting date for the billing period.
    :type start_date: datetime
    :ivar end_date: The ending date for the billing period.
    :type end_date: datetime
    :ivar report: The generated BillingReport object containing summarized billing details for the
        specified parameters.
    :type report: BillingReport
    """
    def __init__(self, customer_id: int, start_date: datetime, end_date: datetime):
        self.customer_id = customer_id
        self.start_date = start_date
        self.end_date = end_date
        self.report = BillingReport(customer_id, start_date, end_date)

    def validate_input(self) -> None:
        """Validate input parameters"""
        try:
            try:
                customer = Customer.objects.get(id=self.customer_id)
            except Customer.DoesNotExist:
                raise ValidationError(f"Customer with ID {self.customer_id} not found")

            if self.start_date > self.end_date:
                raise ValidationError("Start date must be before or equal to end date")

            if not CustomerService.objects.filter(customer_id=self.customer_id).exists():
                raise ValidationError(f"No services found for customer {self.customer_id}")

        except Exception as e:
            logger.error(f"Validation error: {str(e)}")
            raise

    def calculate_service_cost(self, customer_service: CustomerService, order: Order) -> Decimal:
        """
        Calculate the cost for a specific service applied to an order.
        
        Args:
            customer_service: The customer service configuration
            order: The order to calculate service cost for
            
        Returns:
            Decimal: The calculated cost for the service
        """
        try:
            # Check if service should be applied based on rules
            rule_groups = RuleGroup.objects.filter(customer_service=customer_service)
            service_applies = False
            
            if not rule_groups.exists():
                service_applies = True  # If no rules, service always applies
                logger.info(f"No rules found for service {customer_service.service.service_name}, applying by default")
            else:
                for rule_group in rule_groups:
                    if RuleEvaluator.evaluate_rule_group(rule_group, order):
                        service_applies = True
                        logger.info(f"Rule group {rule_group.id} matched for service {customer_service.service.service_name}")
                        break
            
            if not service_applies:
                logger.info(f"Service {customer_service.service.service_name} does not apply to order {order.transaction_id}")
                return Decimal('0')
            
            # Calculate cost based on service type
            if customer_service.service.charge_type == 'single':
                logger.info(f"Applying single charge of {customer_service.unit_price} for service {customer_service.service.service_name}")
                return customer_service.unit_price
            
            # For quantity-based services, calculate based on service type
            service_name = customer_service.service.service_name.lower()
            
            if 'case pick' in service_name or 'case' in service_name:
                total_case_picks = Decimal('0')
                
                # Process each SKU in the order
                for sku, quantity in order.billable_sku_quantities.items():
                    try:
                        product = Product.objects.get(sku=sku, customer_id=order.customer_id)
                        remaining_qty = Decimal(str(quantity))
                        
                        # Get all labeling units and quantities, sorted by quantity in descending order
                        case_sizes = []
                        for i in range(1, 6):
                            unit = getattr(product, f'labeling_unit_{i}')
                            qty = getattr(product, f'labeling_quantity_{i}')
                            if unit and qty and 'case' in unit.lower():
                                case_sizes.append((qty, unit))
                        
                        # Sort by quantity in descending order (largest first)
                        case_sizes.sort(reverse=True)
                        
                        # Calculate cases for each size, starting with largest
                        for case_size, unit_name in case_sizes:
                            cases = remaining_qty // Decimal(str(case_size))
                            if cases > 0:
                                total_case_picks += cases
                                remaining_qty = remaining_qty % Decimal(str(case_size))
                                logger.info(f"Added {cases} {unit_name}(s) of size {case_size} for SKU {sku}")
                    
                    except Product.DoesNotExist:
                        logger.warning(f"Product not found for SKU {sku}")
                        continue
                    except Exception as e:
                        logger.error(f"Error processing case picks for SKU {sku}: {str(e)}")
                        continue
                
                logger.info(f"Calculated total case picks: {total_case_picks} at {customer_service.unit_price} each")
                return customer_service.unit_price * total_case_picks
            
            elif 'pick cost' in service_name or 'pick charge' in service_name:
                total_picks = Decimal('0')
                
                # Process each SKU in the order
                for sku, quantity in order.billable_sku_quantities.items():
                    try:
                        product = Product.objects.get(sku=sku, customer_id=order.customer_id)
                        remaining_qty = Decimal(str(quantity))
                        
                        # Get all case sizes in descending order
                        case_sizes = []
                        for i in range(1, 6):
                            unit = getattr(product, f'labeling_unit_{i}')
                            qty = getattr(product, f'labeling_quantity_{i}')
                            if unit and qty and 'case' in unit.lower():
                                case_sizes.append(qty)
                        
                        # Sort by size descending
                        case_sizes.sort(reverse=True)
                        
                        # Subtract all case quantities from remaining quantity
                        for case_size in case_sizes:
                            cases = remaining_qty // Decimal(str(case_size))
                            if cases > 0:
                                remaining_qty = remaining_qty % Decimal(str(case_size))
                        
                        # Add remaining units as individual picks
                        total_picks += remaining_qty
                        logger.info(f"Added {remaining_qty} individual picks for SKU {sku}")
                        
                    except Product.DoesNotExist:
                        logger.warning(f"Product not found for SKU {sku}")
                        # If product not found, treat all units as individual picks
                        total_picks += Decimal(str(quantity))
                    except Exception as e:
                        logger.error(f"Error processing pick cost for SKU {sku}: {str(e)}")
                        continue
                
                logger.info(f"Calculated total individual picks: {total_picks} at {customer_service.unit_price} each")
                return customer_service.unit_price * total_picks
            
            elif 'sku cost' in service_name:
                # For SKU Cost service, charge based on unique SKUs
                try:
                    sku_data = json.loads(order.sku_quantity) if order.sku_quantity else []
                    unique_skus = len({normalize_sku(item['sku']) for item in sku_data})
                    cost = customer_service.unit_price * Decimal(str(unique_skus))
                    logger.info(f"Calculated SKU Cost: {cost} ({unique_skus} unique SKUs at {customer_service.unit_price} each)")
                    return cost
                except (json.JSONDecodeError, KeyError, AttributeError) as e:
                    logger.error(f"Error calculating SKU Cost: {str(e)}")
                    return Decimal('0')
            else:
                # For other quantity-based services, use total_item_qty
                if order.total_item_qty:
                    cost = customer_service.unit_price * Decimal(str(order.total_item_qty))
                    logger.info(f"Calculated quantity-based cost: {cost} ({order.total_item_qty} items at {customer_service.unit_price} each)")
                    return cost
                else:
                    logger.warning(f"No total_item_qty found for order {order.transaction_id}")
                    return Decimal('0')
            
        except Exception as e:
            logger.error(f"Error calculating service cost: {str(e)}")
            return Decimal('0')

    def generate_report(self) -> BillingReport:
        """
        Generate a billing report for the specified customer and date range.
        
        This method processes all orders within the date range, calculates service costs,
        and aggregates the data into a comprehensive billing report.
        
        Returns:
            BillingReport: A report containing all billing information for the period.
        """
        try:
            self.validate_input()
            
            # Get all completed orders for the customer in the date range
            orders = Order.objects.filter(
                customer_id=self.customer_id,
                close_date__range=(self.start_date, self.end_date)
            ).order_by('close_date')
            
            logger.info(f"Found {orders.count()} orders for customer {self.customer_id}")
            
            # Get all customer services
            customer_services = CustomerService.objects.filter(
                customer_id=self.customer_id
            ).select_related('service')
            
            logger.info(f"Found {customer_services.count()} services for customer {self.customer_id}")
            
            # Initialize service totals dictionary
            self.report.service_totals = {}
            
            # Process each order
            for order in orders:
                logger.info(f"Processing order {order.transaction_id}")
                order_cost = OrderCost(order_id=order.transaction_id)
                
                # Calculate costs for each service
                for cs in customer_services:
                    logger.info(f"Calculating cost for service {cs.service.service_name} on order {order.transaction_id}")
                    service_cost = self.calculate_service_cost(cs, order)
                    logger.info(f"Service {cs.service.service_name} cost: {service_cost}")
                    
                    if service_cost > Decimal('0'):
                        service_cost_obj = ServiceCost(
                            service_id=cs.service.id,
                            service_name=cs.service.service_name,
                            amount=service_cost
                        )
                        order_cost.service_costs.append(service_cost_obj)
                        order_cost.total_amount += service_cost
                        
                        # Update service totals
                        if cs.service.id not in self.report.service_totals:
                            self.report.service_totals[cs.service.id] = {
                                'service_name': cs.service.service_name,
                                'total_amount': Decimal('0')
                            }
                        self.report.service_totals[cs.service.id]['total_amount'] += service_cost
                
                self.report.order_costs.append(order_cost)
                self.report.total_amount += order_cost.total_amount
                logger.info(f"Order {order.transaction_id} total cost: {order_cost.total_amount}")
            
            logger.info(f"Total report amount: {self.report.total_amount}")
            
            # Prepare report data
            self.report.report_data = {
                'customer_id': self.customer_id,
                'start_date': self.start_date.isoformat(),
                'end_date': self.end_date.isoformat(),
                'orders': [
                    {
                        'order_id': oc.order_id,
                        'total_amount': str(oc.total_amount),
                        'services': [
                            {
                                'service_id': sc.service_id,
                                'service_name': sc.service_name,
                                'amount': str(sc.amount)
                            }
                            for sc in oc.service_costs
                        ]
                    }
                    for oc in self.report.order_costs
                ],
                'service_totals': [
                    {
                        'service_id': service_id,
                        'service_name': data['service_name'],
                        'total_amount': str(data['total_amount'])
                    }
                    for service_id, data in self.report.service_totals.items()
                ],
                'total_amount': str(self.report.total_amount)
            }
            
            logger.info(f"Generated report data: {json.dumps(self.report.report_data, indent=2)}")
            return self.report
            
        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")
            raise

    def to_dict(self) -> dict:
        """Convert the report to a dictionary format"""
        try:
            report_dict = {
                'customer_id': self.report.customer_id,
                'start_date': self.report.start_date.isoformat(),
                'end_date': self.report.end_date.isoformat(),
                'order_costs': [
                    {
                        'order_id': order_cost.order_id,
                        'service_costs': [
                            {
                                'service_id': service_cost.service_id,
                                'service_name': service_cost.service_name,
                                'amount': str(service_cost.amount)
                            }
                            for service_cost in order_cost.service_costs
                        ],
                        'total_amount': str(order_cost.total_amount)
                    }
                    for order_cost in self.report.order_costs
                ],
                'service_totals': {
                    str(service_id): str(amount)
                    for service_id, amount in self.report.service_totals.items()
                },
                'total_amount': str(self.report.total_amount)
            }
            self.report.report_data = report_dict
            return report_dict
        except Exception as e:
            logger.error(f"Error converting report to dictionary: {str(e)}")
            raise

    def to_json(self) -> str:
        """
        Convert the report to JSON format
         Convert the report to a dictionary format and then convert it to JSON format using the json.dumps method.
         If an error occurs during the conversion, log the error and return an empty string.
         This method serializes the report data into a JSON string format, which can be easily transmitted or stored.
         It uses the to_dict method to convert the report to a dictionary format before serializing it to JSON.
         The JSON output is formatted with indentation for better readability.

          Returns:
            str: A JSON-formatted string representation of the billing report. The string includes all the details of the report,
            such as customer information, order costs, service totals, and the overall total amount.

          """
        try:
            return json.dumps(self.to_dict(), indent=2)
        except Exception as e:
            logger.error(f"Error converting report to JSON: {str(e)}")
            raise

    def to_csv(self) -> str:
        """Convert the report to CSV format"""
        try:
            lines = [
                "Order ID,Service ID,Service Name,Amount",
            ]

            for order_cost in self.report.order_costs:
                for service_cost in order_cost.service_costs:
                    lines.append(
                        f"{order_cost.order_id},"
                        f"{service_cost.service_id},"
                        f"{service_cost.service_name},"
                        f"{service_cost.amount}"
                    )

            return "\n".join(lines)
        except Exception as e:
            logger.error(f"Error converting report to CSV: {str(e)}")
            raise


def generate_billing_report(
        customer_id: int,
        start_date: Union[datetime, str],
        end_date: Union[datetime, str],
        output_format: str = 'excel'
) -> Union[str, bytes]:
    """
    Generates a billing report for a specified customer over a specified date range 
    and returns the report in the desired format. The report can be generated either as a 
    JSON or CSV file depending on the output format specified.

    :param customer_id: Identifier of the customer for whom the billing report is being generated.
    :type customer_id: int
    :param start_date: Start date of the report. Can be provided as a datetime object or ISO 8601 string.
    :type start_date: Union[datetime, str]
    :param end_date: End date of the report. Can be provided as a datetime object or ISO 8601 string.
    :type end_date: Union[datetime, str]
    :param output_format: Format in which the report will be returned. Options are "excel" (default) or "pdf".
    :type output_format: str
    :return: A string representation of the billing report in the specified format.
    :rtype: Union[str, bytes]
    """
    try:
        logger.info(f"Generating report for customer {customer_id} from {start_date} to {end_date}")

        if isinstance(start_date, str):
            start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        if isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))

        calculator = BillingCalculator(customer_id, start_date, end_date)
        report = calculator.generate_report()

        if output_format.lower() == 'excel':
            return generate_excel_report(report)
        elif output_format.lower() == 'pdf':
            return generate_pdf_report(report)
        return calculator.to_json()

    except Exception as e:
        logger.error(f"Error in generate_billing_report: {str(e)}")
        raise

def generate_excel_report(report: BillingReport) -> bytes:
    """Generate Excel report from billing data"""
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Billing Report Summary"])
    ws_summary.append(["Total Amount", str(report.total_amount)])
    ws_summary.append(["Start Date", report.start_date.strftime('%Y-%m-%d')])
    ws_summary.append(["End Date", report.end_date.strftime('%Y-%m-%d')])
    ws_summary.append([])

    # Service Totals Sheet
    ws_services = wb.create_sheet("Service Totals")
    ws_services.append(["Service", "Amount"])
    for service_id, amount in report.service_totals.items():
        ws_services.append([service_id, str(amount)])

    # Order Details Sheet
    ws_orders = wb.create_sheet("Order Details")
    ws_orders.append(["Order ID", "Service", "Amount", "Total"])
    for order in report.order_costs:
        for service in order.service_costs:
            ws_orders.append([
                order.order_id,
                service.service_name,
                str(service.amount),
                str(order.total_amount)
            ])

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()

def generate_pdf_report(report: BillingReport) -> bytes:
    """Generate PDF report from billing data"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Billing Report", styles['Title']))
    elements.append(Spacer(1, 20))

    # Summary
    summary_data = [
        ["Total Amount:", f"${report.total_amount}"],
        ["Start Date:", report.start_date.strftime('%Y-%m-%d')],
        ["End Date:", report.end_date.strftime('%Y-%m-%d')]
    ]
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Service Totals
    elements.append(Paragraph("Service Totals", styles['Heading2']))
    elements.append(Spacer(1, 10))
    service_data = [["Service", "Amount"]]
    for service_id, amount in report.service_totals.items():
        service_data.append([str(service_id), f"${amount}"])
    service_table = Table(service_data)
    service_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(service_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()